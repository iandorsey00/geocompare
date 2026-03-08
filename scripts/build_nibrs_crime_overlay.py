#!/usr/bin/env python3
"""Build canonical crime overlay metrics from NIBRS agency workbook."""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

from openpyxl import load_workbook

from geocompare.engine import Engine
from geocompare.tools.state_lookup import StateLookup


PLACE_TYPE_SUFFIXES = (
    " city",
    " town",
    " village",
    " borough",
    " municipality",
    " cdp",
    " urban county",
    " unified government",
    " metro government",
)

COUNTY_TYPE_SUFFIXES = (
    " county",
    " parish",
    " borough",
    " census area",
    " municipio",
    " city and borough",
)

COUNTY_AGENCY_TYPES = {
    "county",
    "parish",
    "borough",
    "census area",
    "municipio",
    "metropolitan county",
    "nonmetropolitan county",
}

PLACE_AGENCY_TYPES = {
    "city",
    "town",
    "village",
}

AGENCY_NAME_TRAILING_TERMS = (
    "police department",
    "sheriff s office",
    "sheriffs office",
    "sheriff office",
    "sheriff department",
    "marshal s office",
    "department of public safety",
    "public safety department",
    "department",
    "police",
)

PLACE_ALIAS_MAP = {
    "paso robles": "el paso de robles",
    "ventura": "san buenaventura",
    "lafayette": "la fayette",
}

COUNTY_ALIAS_MAP = {
    ("fl", "jacksonville"): "duval county",
    ("il", "dewitt"): "de witt county",
    ("il", "la salle"): "lasalle county",
    ("la", "la salle"): "lasalle parish",
    ("mt", "anaconda deer lodge"): "deer lodge county",
    ("mt", "butte silver bow"): "silver bow county",
    ("nm", "dona ana"): "dona ana county",
    ("tn", "hartsville trousdale"): "trousdale county",
}


@dataclass
class Agg:
    total: float = 0.0
    persons: float = 0.0
    property: float = 0.0
    population_max: float = 0.0
    rows: int = 0

    def add(self, total: float, persons: float, prop: float, population: float) -> None:
        self.total += total
        self.persons += persons
        self.property += prop
        self.population_max = max(self.population_max, population)
        self.rows += 1


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_name(value: str) -> str:
    value = (
        unicodedata.normalize("NFKD", value)
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )
    value = re.sub(r"\([^)]*\)", " ", value)
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", " ", value).strip()
    return re.sub(r"\s+", " ", value)


def _strip_suffix(name: str, suffixes: Iterable[str]) -> str:
    for suffix in suffixes:
        if name.endswith(suffix):
            return name[: -len(suffix)].strip()
    return name


def _to_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _extract_parenthetical_aliases(label: str) -> Iterable[str]:
    for match in re.findall(r"\(([^)]*)\)", label):
        alias = _norm_name(match)
        if alias:
            yield alias


def _strip_trailing_terms(name: str, terms: Iterable[str]) -> str:
    out = name
    changed = True
    while changed:
        changed = False
        for term in terms:
            if out == term:
                out = ""
                changed = True
                break
            suffix = f" {term}"
            if out.endswith(suffix):
                out = out[: -len(suffix)].strip()
                changed = True
                break
    return out


def _county_name_candidates(state_abbr: str, name: str) -> Iterable[str]:
    base = _norm_name(name)
    if not base:
        return []
    out = {base}
    stripped = _strip_trailing_terms(base, AGENCY_NAME_TRAILING_TERMS)
    if stripped:
        out.add(stripped)
    for candidate in list(out):
        c2 = _strip_suffix(candidate, COUNTY_TYPE_SUFFIXES)
        if c2:
            out.add(c2)
        if c2 and not any(c2.endswith(s) for s in COUNTY_TYPE_SUFFIXES):
            out.add(f"{c2} county")
            out.add(f"{c2} parish")
            out.add(f"{c2} borough")
            out.add(f"{c2} census area")
    for candidate in list(out):
        alias = COUNTY_ALIAS_MAP.get((state_abbr, candidate))
        if alias:
            out.add(alias)
    return [x for x in out if x]


def _place_name_candidates(name: str) -> Iterable[str]:
    base = _norm_name(name)
    if not base:
        return []
    out = {base}
    stripped = _strip_trailing_terms(base, AGENCY_NAME_TRAILING_TERMS)
    if stripped:
        out.add(stripped)
    for candidate in list(out):
        c2 = _strip_suffix(candidate, PLACE_TYPE_SUFFIXES)
        if c2:
            out.add(c2)
        if c2 in PLACE_ALIAS_MAP:
            out.add(PLACE_ALIAS_MAP[c2])
    return [x for x in out if x]


def _build_geo_indexes() -> Tuple[Dict[str, str], Dict[Tuple[str, str], str], Dict[Tuple[str, str], str]]:
    engine = Engine()
    dps = engine.get_data_products()["demographicprofiles"]
    st = StateLookup()

    state_index: Dict[str, str] = {}
    county_index: Dict[Tuple[str, str], str] = {}
    place_index: Dict[Tuple[str, str], str] = {}

    for dp in dps:
        if dp.sumlevel == "040":
            state_name = dp.name.split(",")[0]
            state_index[_norm_name(state_name)] = dp.geoid
            continue

        state_abbr = dp.state.lower()
        label = dp.name.split(",")[0]
        normalized = _norm_name(label)

        if dp.sumlevel == "050":
            county_index[(state_abbr, normalized)] = dp.geoid
            stripped = _norm_name(_strip_suffix(normalized, COUNTY_TYPE_SUFFIXES))
            county_index[(state_abbr, stripped)] = dp.geoid
            for alias in _extract_parenthetical_aliases(label):
                county_index[(state_abbr, alias)] = dp.geoid
            continue

        if dp.sumlevel == "160":
            place_index[(state_abbr, normalized)] = dp.geoid
            stripped = _norm_name(_strip_suffix(normalized, PLACE_TYPE_SUFFIXES))
            place_index[(state_abbr, stripped)] = dp.geoid
            for alias in _extract_parenthetical_aliases(label):
                place_index[(state_abbr, alias)] = dp.geoid

    # Make sure we can map state names in workbook form.
    for geoid, state_name in st.geoid_to_name.items():
        key = _norm_name(state_name)
        if key not in state_index:
            state_index[key] = f"04000US{geoid}"

    return state_index, county_index, place_index


def _parse_workbook(path: Path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    # Header is at row 5 in this report format.
    for row in ws.iter_rows(min_row=7, values_only=True):
        state_name = _clean_text(row[0])
        agency_type = _clean_text(row[1])
        agency_name = _clean_text(row[2])
        if not state_name or not agency_type or not agency_name:
            continue
        if state_name.lower() in {"united states", "total"}:
            continue

        yield {
            "state_name": state_name,
            "agency_type": agency_type,
            "agency_name": agency_name,
            "population": _to_float(row[3]),
            "total_offenses": _to_float(row[4]),
            "crimes_against_persons": _to_float(row[5]),
            "crimes_against_property": _to_float(row[6]),
        }


def _write_crime_overlay(out_path: Path, rows: Dict[str, Agg]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["GEOID", "violent_crime_count", "property_crime_count", "total_crime_count"]
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for geoid in sorted(rows):
            agg = rows[geoid]
            writer.writerow(
                {
                    "GEOID": geoid,
                    # NIBRS "Crimes Against Persons" is used as the violent proxy.
                    "violent_crime_count": int(round(agg.persons)),
                    "property_crime_count": int(round(agg.property)),
                    "total_crime_count": int(round(agg.total)),
                }
            )


def _write_coverage_report(out_path: Path, report: dict) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(report, indent=2, sort_keys=True), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Build base crime overlay from NIBRS workbook.")
    parser.add_argument(
        "--nibrs-xlsx",
        default="/Users/iandorsey/Downloads/NIBRS_United_States_Offense_Type_by_Agency_2024.xlsx",
        help="path to NIBRS offense-by-agency workbook",
    )
    parser.add_argument(
        "--out-dir",
        default="/Users/iandorsey/dev/000-data",
        help="geocompare data directory",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.nibrs_xlsx).expanduser()
    out_dir = Path(args.out_dir).expanduser()
    overlay_path = out_dir / "overlays" / "crime_data.csv"
    report_path = out_dir / "overlays" / "crime_data_coverage.json"

    state_index, county_index, place_index = _build_geo_indexes()

    state_aggs: Dict[str, Agg] = defaultdict(Agg)
    county_aggs: Dict[str, Agg] = defaultdict(Agg)
    place_aggs: Dict[str, Agg] = defaultdict(Agg)

    coverage = {
        "input_rows": 0,
        "mapped": {"state": 0, "county": 0, "place": 0},
        "unmapped": {"state": 0, "county": 0, "place": 0},
        "unmapped_examples": {"state": [], "county": [], "place": []},
    }

    st = StateLookup()
    state_name_to_abbrev_ci = {name.lower(): abbr.lower() for name, abbr in st.name_to_abbrev.items()}

    for rec in _parse_workbook(xlsx_path):
        coverage["input_rows"] += 1
        state_name = rec["state_name"]
        state_norm = _norm_name(state_name)
        state_geoid = state_index.get(state_norm)
        state_abbr = state_name_to_abbrev_ci.get(state_name.lower(), "")
        name_norm = _norm_name(rec["agency_name"])
        agency_type_norm = rec["agency_type"].strip().lower()

        if state_geoid:
            state_aggs[state_geoid].add(
                rec["total_offenses"],
                rec["crimes_against_persons"],
                rec["crimes_against_property"],
                rec["population"],
            )
            coverage["mapped"]["state"] += 1
        else:
            coverage["unmapped"]["state"] += 1
            if len(coverage["unmapped_examples"]["state"]) < 15:
                coverage["unmapped_examples"]["state"].append(state_name)

        if agency_type_norm in COUNTY_AGENCY_TYPES:
            county_geoid: Optional[str] = None
            for candidate in _county_name_candidates(state_abbr, rec["agency_name"]):
                county_geoid = county_index.get((state_abbr, candidate))
                if county_geoid:
                    break
            if county_geoid:
                county_aggs[county_geoid].add(
                    rec["total_offenses"],
                    rec["crimes_against_persons"],
                    rec["crimes_against_property"],
                    rec["population"],
                )
                coverage["mapped"]["county"] += 1
            else:
                coverage["unmapped"]["county"] += 1
                if len(coverage["unmapped_examples"]["county"]) < 20:
                    coverage["unmapped_examples"]["county"].append(
                        f"{rec['agency_name']} ({state_name})"
                    )
            continue

        if agency_type_norm in PLACE_AGENCY_TYPES:
            place_geoid: Optional[str] = None
            for candidate in _place_name_candidates(rec["agency_name"]):
                place_geoid = place_index.get((state_abbr, candidate))
                if place_geoid:
                    break
            if place_geoid:
                place_aggs[place_geoid].add(
                    rec["total_offenses"],
                    rec["crimes_against_persons"],
                    rec["crimes_against_property"],
                    rec["population"],
                )
                coverage["mapped"]["place"] += 1
            else:
                coverage["unmapped"]["place"] += 1
                if len(coverage["unmapped_examples"]["place"]) < 20:
                    coverage["unmapped_examples"]["place"].append(
                        f"{rec['agency_name']} ({state_name})"
                    )

    all_rows: Dict[str, Agg] = {}
    all_rows.update(state_aggs)
    all_rows.update(county_aggs)
    all_rows.update(place_aggs)
    _write_crime_overlay(overlay_path, all_rows)

    coverage["output_rows"] = {
        "state": len(state_aggs),
        "county": len(county_aggs),
        "place": len(place_aggs),
        "total": len(all_rows),
    }
    _write_coverage_report(report_path, coverage)

    print(f"wrote {overlay_path} ({len(all_rows)} rows)")
    print(f"coverage report: {report_path}")
    print(json.dumps(coverage["output_rows"], indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
